from __future__ import annotations

import csv
import shutil
import subprocess
from pathlib import Path

import pytest

from tools import build_private_label_review_workbook as workbook_builder


ROOT = Path(__file__).resolve().parents[1]
RESTRICTED_TEST_DIR = ROOT / "data" / "restricted" / "private_whatsapp" / "processed" / "pytest_workbook"
SYNTHETIC_RAW_TEXT = "SYNTHETIC REVIEW TEXT MUST NOT PRINT"
FIELDNAMES = [
    "example_id",
    "split",
    "speaker_roles",
    "text_window_redacted",
    "candidate_labels",
    "evidence_hint",
    "review_label",
    "severity",
    "safe_next_step",
    "reviewer_notes",
]


def _cleanup() -> None:
    shutil.rmtree(RESTRICTED_TEST_DIR, ignore_errors=True)


def _write_synthetic_csv(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerow(
            {
                "example_id": "synthetic_001",
                "split": "synthetic",
                "speaker_roles": "self|other",
                "text_window_redacted": SYNTHETIC_RAW_TEXT,
                "candidate_labels": "direct_ask;unclear_timing",
                "evidence_hint": "synthetic hint",
                "review_label": "",
                "severity": "",
                "safe_next_step": "",
                "reviewer_notes": "",
            }
        )
    return path


def test_script_refuses_non_restricted_input_output(tmp_path: Path) -> None:
    _cleanup()
    try:
        restricted_input = _write_synthetic_csv(RESTRICTED_TEST_DIR / "review.csv")
        restricted_output = RESTRICTED_TEST_DIR / "review.xlsx"
        outside_input = _write_synthetic_csv(tmp_path / "review.csv")

        assert workbook_builder.main(["--input", str(outside_input), "--output", str(restricted_output)]) == 2
        assert workbook_builder.main(["--input", str(restricted_input), "--output", str(tmp_path / "review.xlsx")]) == 2
    finally:
        _cleanup()


def test_no_raw_private_text_is_printed(capsys) -> None:
    _cleanup()
    try:
        restricted_input = _write_synthetic_csv(RESTRICTED_TEST_DIR / "review.csv")
        restricted_output = RESTRICTED_TEST_DIR / "review.xlsx"

        workbook_builder.main(["--input", str(restricted_input), "--output", str(restricted_output)])

        captured = capsys.readouterr()
        assert SYNTHETIC_RAW_TEXT not in captured.out
        assert SYNTHETIC_RAW_TEXT not in captured.err
    finally:
        _cleanup()


def test_workbook_created_with_sheets_and_dropdowns(capsys) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.worksheet.datavalidation import DataValidation

    _cleanup()
    try:
        restricted_input = _write_synthetic_csv(RESTRICTED_TEST_DIR / "review.csv")
        restricted_output = RESTRICTED_TEST_DIR / "review.xlsx"

        exit_code = workbook_builder.main(["--input", str(restricted_input), "--output", str(restricted_output)])

        captured = capsys.readouterr()
        assert exit_code == 0
        assert SYNTHETIC_RAW_TEXT not in captured.out
        assert SYNTHETIC_RAW_TEXT not in captured.err
        assert restricted_output.exists()

        ignored = subprocess.run(["git", "check-ignore", "-q", str(restricted_output.relative_to(ROOT))], cwd=ROOT, check=False)
        assert ignored.returncode == 0

        wb = openpyxl.load_workbook(restricted_output)
        assert {"Review", "Guide", "Choices"} <= set(wb.sheetnames)
        review = wb["Review"]
        assert [cell.value for cell in review[1]] == FIELDNAMES

        validations = list(review.data_validations.dataValidation)
        assert _has_dropdown_for_header(review, validations, "review_label")
        assert _has_dropdown_for_header(review, validations, "severity")
        assert _has_dropdown_for_header(review, validations, "safe_next_step")
        assert not _has_dropdown_for_header(review, validations, "reviewer_notes")
    finally:
        _cleanup()


def _has_dropdown_for_header(sheet, validations: list[DataValidation], header: str) -> bool:
    headers = [cell.value for cell in sheet[1]]
    column_index = headers.index(header) + 1
    letter = sheet.cell(row=1, column=column_index).column_letter
    return any(dv.type == "list" and f"{letter}2" in str(dv.sqref) for dv in validations)
