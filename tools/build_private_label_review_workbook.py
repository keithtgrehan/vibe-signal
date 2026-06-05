#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
RESTRICTED_ROOT = REPO_ROOT / "data" / "restricted" / "private_whatsapp"
DEFAULT_INPUT = RESTRICTED_ROOT / "processed" / "private_label_review_100.csv"
DEFAULT_OUTPUT = RESTRICTED_ROOT / "processed" / "private_label_review_100.xlsx"

REVIEW_LABELS = [
    "unclear_timing",
    "direct_ask",
    "unanswered_ask_candidate",
    "pressure_urgency",
    "boundary",
    "boundary_pressure",
    "reassurance",
    "repair_attempt",
    "escalation_risk",
    "ambiguity",
    "cognitive_load",
    "specificity_drop",
    "soft_commitment",
    "low_signal",
    "none",
]
SEVERITIES = ["low", "medium", "high"]
SAFE_NEXT_STEPS = [
    "unclear_timing: Ask for a specific confirmation time.",
    "direct_ask: Answer the ask directly or ask one clarifying question.",
    "unanswered_ask_candidate: Restate the question once in simpler terms.",
    "pressure_urgency: Slow the exchange down and respond with one clear boundary or timeframe.",
    "boundary: Respect the boundary and confirm when to reconnect.",
    "boundary_pressure: Restate the boundary calmly and avoid debating it.",
    "reassurance: Acknowledge the reassurance need clearly.",
    "repair_attempt: Acknowledge the repair before reopening the issue.",
    "escalation_risk: Pause and respond to the concrete issue, not the tone.",
    "ambiguity: Ask what they mean instead of interpreting it.",
    "cognitive_load: Split the reply into one issue at a time.",
    "specificity_drop: Ask whether the plan has changed.",
    "soft_commitment: Ask for confirmation before relying on it.",
    "low_signal: No action needed; ask for clarification only if it matters.",
    "none: Reject candidate label; no safe next step needed.",
]
GUIDE_ROWS = [
    {
        "label": "unclear_timing",
        "definition": "Timing language is vague or leaves the next time unclear.",
        "when_to_use": "Use when a message points to later, soon, maybe, or another unspecific time.",
        "default_safe_next_step": "Ask for a specific confirmation time.",
        "what_not_to_infer": "Do not infer avoidance, intent, or relationship priority.",
    },
    {
        "label": "direct_ask",
        "definition": "The text contains a direct request or question.",
        "when_to_use": "Use when a concrete answer or action is being requested.",
        "default_safe_next_step": "Answer the ask directly or ask one clarifying question.",
        "what_not_to_infer": "Do not infer pressure unless pressure wording is present.",
    },
    {
        "label": "unanswered_ask_candidate",
        "definition": "A prior ask may not have received a clear answer in the local window.",
        "when_to_use": "Use when the reply shifts topic, stays vague, or lacks an answer marker.",
        "default_safe_next_step": "Restate the question once in simpler terms.",
        "what_not_to_infer": "Do not infer ignoring, deception, or hidden motive.",
    },
    {
        "label": "pressure_urgency",
        "definition": "The wording adds urgency, obligation, or constrained response time.",
        "when_to_use": "Use for terms like urgent, right now, must, or have to.",
        "default_safe_next_step": "Slow the exchange down and respond with one clear boundary or timeframe.",
        "what_not_to_infer": "Do not label a person as controlling or manipulative.",
    },
    {
        "label": "boundary",
        "definition": "The text states a limit, refusal, availability constraint, or preference boundary.",
        "when_to_use": "Use when a person says no, cannot, is unavailable, or needs space.",
        "default_safe_next_step": "Respect the boundary and confirm when to reconnect.",
        "what_not_to_infer": "Do not infer rejection or relationship health.",
    },
    {
        "label": "boundary_pressure",
        "definition": "A boundary is followed by wording that pressures against it.",
        "when_to_use": "Use when a refusal or limit is met with obligation, urgency, or debate pressure.",
        "default_safe_next_step": "Restate the boundary calmly and avoid debating it.",
        "what_not_to_infer": "Do not infer abuse, intent, or personality traits.",
    },
    {
        "label": "reassurance",
        "definition": "The text uses pressure-lowering or reassurance wording.",
        "when_to_use": "Use for no pressure, no rush, all good, or similar language.",
        "default_safe_next_step": "Acknowledge the reassurance need clearly.",
        "what_not_to_infer": "Do not infer attachment style, anxiety, or affection.",
    },
    {
        "label": "repair_attempt",
        "definition": "The text attempts to reset, apologize, clarify, or repair the exchange.",
        "when_to_use": "Use when wording includes sorry, reset, rephrase, or clarify after friction.",
        "default_safe_next_step": "Acknowledge the repair before reopening the issue.",
        "what_not_to_infer": "Do not infer remorse as an internal state.",
    },
    {
        "label": "escalation_risk",
        "definition": "The exchange contains observable intensity markers that can make replies harder.",
        "when_to_use": "Use for repeated exclamation, blame wording, or broad conflict phrasing.",
        "default_safe_next_step": "Pause and respond to the concrete issue, not the tone.",
        "what_not_to_infer": "Do not infer anger, instability, or hidden emotion.",
    },
    {
        "label": "ambiguity",
        "definition": "The wording is underspecified or open to multiple readings.",
        "when_to_use": "Use when meaning is unclear without more context.",
        "default_safe_next_step": "Ask what they mean instead of interpreting it.",
        "what_not_to_infer": "Do not infer secret intent.",
    },
    {
        "label": "cognitive_load",
        "definition": "The message contains many requests, topics, or decisions at once.",
        "when_to_use": "Use when a reply would need to handle multiple items together.",
        "default_safe_next_step": "Split the reply into one issue at a time.",
        "what_not_to_infer": "Do not infer overwhelm, neurotype, or mental state.",
    },
    {
        "label": "specificity_drop",
        "definition": "Earlier concrete wording becomes less specific in a later local window.",
        "when_to_use": "Use when a concrete plan turns into maybe, later, or unclear details.",
        "default_safe_next_step": "Ask whether the plan has changed.",
        "what_not_to_infer": "Do not infer deception or changed feelings.",
    },
    {
        "label": "soft_commitment",
        "definition": "The message signals possible agreement without a firm commitment.",
        "when_to_use": "Use for maybe, should, probably, or tentative plan wording.",
        "default_safe_next_step": "Ask for confirmation before relying on it.",
        "what_not_to_infer": "Do not infer unreliability or intent.",
    },
    {
        "label": "low_signal",
        "definition": "The window has too little context or too few cues for a useful label.",
        "when_to_use": "Use for very short, generic, or context-light windows.",
        "default_safe_next_step": "No action needed; ask for clarification only if it matters.",
        "what_not_to_infer": "Do not infer emotion, intent, or relationship state.",
    },
    {
        "label": "none",
        "definition": "The candidate label is not supported by the observable window.",
        "when_to_use": "Use when no listed cue applies.",
        "default_safe_next_step": "Reject candidate label; no safe next step needed.",
        "what_not_to_infer": "Do not replace unsupported labels with guesses.",
    },
]
WRAP_COLUMNS = {"text_window_redacted", "evidence_hint", "safe_next_step", "reviewer_notes"}
WIDTHS = {
    "example_id": 18,
    "split": 16,
    "speaker_roles": 18,
    "text_window_redacted": 54,
    "candidate_labels": 32,
    "evidence_hint": 36,
    "review_label": 24,
    "severity": 12,
    "safe_next_step": 58,
    "reviewer_notes": 44,
}


def _resolve(path: Path) -> Path:
    return path.expanduser().resolve()


def _is_under(path: Path, root: Path) -> bool:
    try:
        _resolve(path).relative_to(_resolve(root))
        return True
    except ValueError:
        return False


def _ensure_restricted(path: Path, *, kind: str) -> Path:
    resolved = _resolve(path)
    if not _is_under(resolved, RESTRICTED_ROOT):
        raise ValueError(f"{kind} must be under data/restricted/private_whatsapp")
    return resolved


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        return fieldnames, [dict(row) for row in reader]


def _load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Alignment
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build the workbook. Install it with: pip install openpyxl") from exc
    return Workbook, Font, PatternFill, get_column_letter, DataValidation, Alignment


def _choices_formula(sheet_name: str, start_row: int, end_row: int) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!$A${start_row}:$A${end_row}"


def build_workbook(input_csv: Path, output_xlsx: Path) -> dict[str, Any]:
    safe_input = _ensure_restricted(input_csv, kind="input CSV")
    safe_output = _ensure_restricted(output_xlsx, kind="output XLSX")
    if safe_output.suffix.lower() != ".xlsx":
        raise ValueError("output path must end with .xlsx")
    if not safe_input.exists():
        raise FileNotFoundError("input CSV not found")

    fieldnames, rows = _read_csv(safe_input)
    if not fieldnames:
        raise ValueError("input CSV must contain a header row")

    Workbook, Font, PatternFill, get_column_letter, DataValidation, Alignment = _load_openpyxl()
    workbook = Workbook()
    review = workbook.active
    review.title = "Review"
    guide = workbook.create_sheet("Guide")
    choices = workbook.create_sheet("Choices")

    _populate_review_sheet(review, fieldnames, rows, Font, PatternFill, Alignment, get_column_letter)
    _populate_guide_sheet(guide, Font, PatternFill, Alignment)
    choice_ranges = _populate_choices_sheet(choices, Font, PatternFill)
    choices.sheet_state = "hidden"
    _add_dropdowns(review, fieldnames, len(rows), choice_ranges, DataValidation)

    safe_output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(safe_output)
    return {"output": str(safe_output), "row_count": len(rows), "dropdowns_added": ["review_label", "severity", "safe_next_step"]}


def _populate_review_sheet(sheet, fieldnames, rows, Font, PatternFill, Alignment, get_column_letter) -> None:
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, field in enumerate(fieldnames, start=1):
        width = WIDTHS.get(field, min(max(len(field) + 4, 14), 28))
        sheet.column_dimensions[get_column_letter(index)].width = width
        if field in WRAP_COLUMNS:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2, max_row=max(2, sheet.max_row)):
                for item in cell:
                    item.alignment = Alignment(wrap_text=True, vertical="top")


def _populate_guide_sheet(sheet, Font, PatternFill, Alignment) -> None:
    headers = ["label", "plain_english_definition", "when_to_use", "default_safe_next_step", "what_not_to_infer"]
    sheet.append(headers)
    for row in GUIDE_ROWS:
        sheet.append([row["label"], row["definition"], row["when_to_use"], row["default_safe_next_step"], row["what_not_to_infer"]])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    widths = [24, 42, 48, 52, 46]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=1, max_row=sheet.max_row):
            for item in cell:
                item.alignment = Alignment(wrap_text=True, vertical="top")


def _populate_choices_sheet(sheet, Font, PatternFill) -> dict[str, str]:
    groups = {
        "review_label": REVIEW_LABELS,
        "severity": SEVERITIES,
        "safe_next_step": SAFE_NEXT_STEPS,
    }
    current_col = 1
    ranges: dict[str, str] = {}
    header_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for name, values in groups.items():
        sheet.cell(row=1, column=current_col).value = name
        sheet.cell(row=1, column=current_col).font = Font(bold=True)
        sheet.cell(row=1, column=current_col).fill = header_fill
        for row_index, value in enumerate(values, start=2):
            sheet.cell(row=row_index, column=current_col).value = value
        col_letter = sheet.cell(row=1, column=current_col).column_letter
        ranges[name] = f"'Choices'!${col_letter}$2:${col_letter}${len(values) + 1}"
        sheet.column_dimensions[col_letter].width = 72 if name == "safe_next_step" else 28
        current_col += 1
    return ranges


def _add_dropdowns(sheet, fieldnames: list[str], row_count: int, choice_ranges: dict[str, str], DataValidation) -> None:
    max_row = max(row_count + 1, 2)
    for field in ("review_label", "severity", "safe_next_step"):
        if field not in fieldnames:
            continue
        column_index = fieldnames.index(field) + 1
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        validation = DataValidation(type="list", formula1=choice_ranges[field], allow_blank=True)
        validation.error = "Choose a value from the dropdown list."
        validation.errorTitle = "Invalid choice"
        validation.prompt = "Select a safe review value."
        validation.promptTitle = field
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{max_row}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build a local-only XLSX workbook for private WhatsApp label review.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args(argv)

    try:
        summary = build_workbook(Path(args.input), Path(args.output))
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2

    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
