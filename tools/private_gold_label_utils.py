#!/usr/bin/env python3
from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
RESTRICTED_ROOT = REPO_ROOT / "data" / "restricted" / "private_whatsapp"

ROW_ID_COLUMNS = ("example_id", "row_id", "label_id")
GOLD_LABEL_COLUMNS = ("review_label", "primary_cue", "cue_id")
PREDICTION_COLUMNS = ("deterministic_label", "weak_label", "predicted_label", "candidate_labels")
REVIEW_STATUS_COLUMNS = ("review_status", "status")
SEVERITY_COLUMNS = ("severity",)
SAFE_NEXT_STEP_COLUMNS = ("safe_next_step", "safe_next_step_id")
EVIDENCE_COLUMNS = (
    "evidence_span_start",
    "evidence_span_end",
    "evidence_start",
    "evidence_end",
    "evidence_text",
    "evidence_hint",
    "evidence_span_text_local",
)

KNOWN_COLUMNS = {
    *ROW_ID_COLUMNS,
    *GOLD_LABEL_COLUMNS,
    *PREDICTION_COLUMNS,
    *REVIEW_STATUS_COLUMNS,
    *SEVERITY_COLUMNS,
    *SAFE_NEXT_STEP_COLUMNS,
    *EVIDENCE_COLUMNS,
    "split",
    "speaker_roles",
    "text_window_redacted",
    "source_id",
    "secondary_cues",
    "confidence",
    "reviewer_confidence",
    "reviewer_notes",
    "review_notes_local",
    "cannot_infer_flags",
    "unsafe_model_output",
    "needs_more_context",
    "reviewer_id",
    "reviewed_at",
}

ALLOWED_LABELS = {
    "unclear_timing",
    "direct_ask",
    "unanswered_ask_candidate",
    "pressure_urgency",
    "boundary",
    "boundary_pressure",
    "reassurance",
    "repair_attempt",
    "escalation_risk",
    "ambiguity",
    "cognitive_load",
    "specificity_drop",
    "soft_commitment",
    "low_signal",
    "none",
    "clarity",
    "pressure",
    "directness",
    "repair_opportunity",
    "safety_boundary",
    "reject",
    "invalid",
    "needs_more_context",
}

REJECT_LABELS = {"none", "reject", "invalid"}
LOW_SIGNAL_LABELS = {"low_signal"}
ALLOWED_SEVERITIES = {"low", "medium", "high", "not_applicable"}
SAFE_NEXT_STEP_IDS = {
    "ask_clear_follow_up",
    "offer_time_window",
    "reduce_pressure",
    "acknowledge_boundary",
    "name_uncertainty",
    "repair_and_reset",
    "pause_and_wait",
    "choose_low_stakes_check_in",
    "do_not_send_coercive_reply",
    "not_enough_context",
}


class PrivateGoldLabelError(ValueError):
    """Safe validation error. Message must not contain private row content."""


def _resolve(path: Path) -> Path:
    return path.expanduser().resolve()


def ensure_restricted_path(path: Path, *, kind: str, restricted_root: Path = RESTRICTED_ROOT) -> Path:
    resolved = _resolve(path)
    root = _resolve(restricted_root)
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise PrivateGoldLabelError(f"{kind} must be under restricted private root") from exc
    return resolved


def ensure_restricted_report_path(path: Path, *, kind: str, restricted_root: Path = RESTRICTED_ROOT) -> Path:
    resolved = ensure_restricted_path(path, kind=kind, restricted_root=restricted_root)
    reports_root = _resolve(restricted_root) / "reports"
    try:
        resolved.relative_to(reports_root)
    except ValueError as exc:
        raise PrivateGoldLabelError(f"{kind} must be under restricted private reports root") from exc
    return resolved


def infer_format(path: Path, requested: str | None = None) -> str:
    if requested:
        return requested.lower()
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    raise PrivateGoldLabelError("unsupported input format")


def read_table(path: Path, *, file_format: str | None = None, restricted_root: Path = RESTRICTED_ROOT) -> tuple[list[str], list[dict[str, str]]]:
    safe_path = ensure_restricted_path(path, kind="input", restricted_root=restricted_root)
    if not safe_path.exists():
        raise PrivateGoldLabelError("input file not found")
    resolved_format = infer_format(safe_path, file_format)
    if resolved_format == "csv":
        return read_csv(safe_path)
    if resolved_format == "xlsx":
        return read_xlsx(safe_path)
    raise PrivateGoldLabelError("unsupported input format")


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        return fieldnames, [{key: str(value or "") for key, value in row.items()} for row in reader]


def read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise PrivateGoldLabelError("xlsx support requires openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    fieldnames = [str(value or "").strip() for value in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        records.append({field: str(value or "") for field, value in zip(fieldnames, row)})
    return fieldnames, records


def first_present(fieldnames: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {field.lower(): field for field in fieldnames}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None


def split_labels(value: str) -> list[str]:
    normalized = value.replace("|", ";").replace(",", ";")
    labels = [item.strip().lower() for item in normalized.split(";") if item.strip()]
    return labels


def normalize_label_set(value: str) -> set[str]:
    return set(split_labels(value))


def categorize_safe_next_step(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return "blank"
    if normalized in SAFE_NEXT_STEP_IDS:
        return normalized
    prefix = normalized.split(":", 1)[0].strip()
    if prefix in ALLOWED_LABELS:
        return prefix
    return "custom_or_unrecognized"


def normalize_severity(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return "blank"
    if normalized in ALLOWED_SEVERITIES:
        return normalized
    return "custom_or_unrecognized"


def has_any_value(row: dict[str, str], columns: tuple[str, ...] | list[str]) -> bool:
    return any(str(row.get(column, "")).strip() for column in columns)


def safe_counter(counter: Counter[str]) -> dict[str, int]:
    return dict(sorted(counter.items()))
